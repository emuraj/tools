"""
SUMMARY
-------
This script opens an Excel workbook, scans all worksheets and all text cells,
and removes:
1. A leading "1)" only when it appears at the start of a cell
2. All occurrences of "**" anywhere in the cell

The cleaned workbook is then saved as a NEW Excel file in the SAME folder
as the original workbook, with "_cleaned" added to the filename.

Example:
    Original: C:\Folder\INV_Summaries.xlsx
    Output:   C:\Folder\INV_Summaries_cleaned.xlsx
"""

import re
from pathlib import Path
from openpyxl import load_workbook


# =========================================================
# INPUT EXCEL FILE PATH  <<< PUT YOUR FILE PATH HERE
# =========================================================
EXCEL_FILE = r"C:\Users\e.muraj\OneDrive - Neurotech USA, Inc\production26-csv\record_title_synopsis_20260301_135052.xlsx"
# =========================================================


def clean_text(value):
    """Remove leading '1)' and all '**' from text cells."""
    if not isinstance(value, str):
        return value

    text = value

    # Remove leading 1) only if it appears at the start of the cell
    text = re.sub(r'^\s*1\)\s*', '', text)

    # Remove all markdown bold markers
    text = text.replace("**", "")

    return text


def main():
    input_path = Path(EXCEL_FILE)

    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")

    # Output goes in the SAME folder as the original file
    output_path = input_path.with_name(f"{input_path.stem}_cleaned{input_path.suffix}")

    wb = load_workbook(input_path)
    changed_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    original = cell.value
                    cleaned = clean_text(original)

                    if cleaned != original:
                        cell.value = cleaned
                        changed_cells += 1

    wb.save(output_path)

    print("Done.")
    print(f"Original file: {input_path}")
    print(f"Cleaned file : {output_path}")
    print(f"Cells changed: {changed_cells}")


if __name__ == "__main__":
    main()