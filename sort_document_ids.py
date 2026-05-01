# sort_document_ids_to_xlsx.py
"""
Purpose:
    Sort a CSV by Document ID and produce an XLSX workbook with review tabs.

What this file does:
    Reads an input CSV file, sorts rows by Document ID using natural numeric ordering,
    identifies Document IDs with spaced dash formatting, identifies normalized duplicate
    IDs that exist in both spaced and unspaced forms, and writes the results to an XLSX
    workbook with multiple worksheets.

Place in the larger scheme:
    This can be used as a data-quality cleanup and review step before validating,
    importing, reconciling, or remediating controlled document records.
"""

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# FILE PATHS — EDIT THESE VALUES
# =============================================================================

INPUT_CSV_PATH = Path(
    r"C:\Users\e.muraj\Downloads\Intellect_20260427_074117635_100000102000000080.csv"
)

OUTPUT_XLSX_FOLDER_PATH = Path(r"C:\Users\e.muraj\Downloads")

OUTPUT_XLSX_FILE_NAME = "sorted_document_id_review.xlsx"

DOCUMENT_ID_COLUMN_NAME = "Document ID"

LEGACY_DOCUMENT_NUMBER_COLUMN_LETTER = "AO"

# =============================================================================
# END FILE PATHS
# =============================================================================


DOCUMENT_ID_PATTERN = re.compile(r"^\s*([A-Za-z]+)\s*-\s*(\d+)\s*$")
SPACED_DASH_PATTERN = re.compile(r"\s-\s|\s-|-\s")


def column_letter_to_zero_based_index(column_letter: str) -> int:
    index = 0

    for character in column_letter.upper():
        index = index * 26 + (ord(character) - ord("A") + 1)

    return index - 1


def get_cell_value(row: list[str], index: int) -> str:
    if index < len(row):
        return row[index]

    return ""


def normalize_document_id(document_id: str) -> str:
    match = DOCUMENT_ID_PATTERN.match(document_id.strip())

    if not match:
        return document_id.strip().upper()

    prefix = match.group(1).upper()
    number = int(match.group(2))

    return f"{prefix}-{number}"


def document_id_sort_key(document_id: str) -> tuple:
    match = DOCUMENT_ID_PATTERN.match(document_id.strip())

    if not match:
        return ("ZZZZZZZZZZ", float("inf"), document_id.strip().upper())

    prefix = match.group(1).upper()
    number = int(match.group(2))

    return (prefix, number, document_id.strip().upper())


def has_spaced_dash(document_id: str) -> bool:
    return bool(SPACED_DASH_PATTERN.search(document_id))


def build_review_header(header: list[str], document_id_index: int, legacy_document_index: int) -> list[str]:
    legacy_header = get_cell_value(header, legacy_document_index) or "Legacy Document #"

    review_header = [
        get_cell_value(header, document_id_index),
        legacy_header,
    ]

    for index, column_name in enumerate(header):
        if index not in {document_id_index, legacy_document_index}:
            review_header.append(column_name)

    return review_header


def build_review_row(row: list[str], document_id_index: int, legacy_document_index: int, header_length: int) -> list[str]:
    normalized_row = row + [""] * max(0, header_length - len(row))

    review_row = [
        get_cell_value(normalized_row, document_id_index),
        get_cell_value(normalized_row, legacy_document_index),
    ]

    for index, value in enumerate(normalized_row[:header_length]):
        if index not in {document_id_index, legacy_document_index}:
            review_row.append(value)

    return review_row


def style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_border = Border(
        bottom=Side(style="thin", color="B7B7B7")
    )

    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        adjusted_width = min(max(max_length + 2, 12), 45)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    worksheet.auto_filter.ref = worksheet.dimensions


def write_rows_to_sheet(worksheet, rows: list[list[str]]) -> None:
    for row in rows:
        worksheet.append(row)


def main() -> None:
    if not INPUT_CSV_PATH.exists():
        raise FileNotFoundError(f"Input CSV not found: {INPUT_CSV_PATH}")

    if INPUT_CSV_PATH.is_dir():
        raise IsADirectoryError(f"Input path is a folder, not a CSV file: {INPUT_CSV_PATH}")

    OUTPUT_XLSX_FOLDER_PATH.mkdir(parents=True, exist_ok=True)

    if not OUTPUT_XLSX_FOLDER_PATH.is_dir():
        raise NotADirectoryError(f"Output path is not a folder: {OUTPUT_XLSX_FOLDER_PATH}")

    output_xlsx_path = OUTPUT_XLSX_FOLDER_PATH / OUTPUT_XLSX_FILE_NAME

    with INPUT_CSV_PATH.open("r", newline="", encoding="utf-8-sig") as infile:
        reader = csv.reader(infile)
        header = next(reader, None)

        if header is None:
            raise ValueError("Input CSV is empty.")

        rows = list(reader)

    try:
        document_id_index = header.index(DOCUMENT_ID_COLUMN_NAME)
    except ValueError:
        raise ValueError(f"Column '{DOCUMENT_ID_COLUMN_NAME}' was not found in the first row.")

    legacy_document_index = column_letter_to_zero_based_index(LEGACY_DOCUMENT_NUMBER_COLUMN_LETTER)

    sorted_rows = sorted(
        rows,
        key=lambda row: document_id_sort_key(get_cell_value(row, document_id_index)),
    )

    spaced_dash_rows = [
        row for row in sorted_rows
        if has_spaced_dash(get_cell_value(row, document_id_index))
    ]

    rows_by_normalized_id = defaultdict(list)

    for row in sorted_rows:
        document_id = get_cell_value(row, document_id_index)
        normalized_id = normalize_document_id(document_id)
        rows_by_normalized_id[normalized_id].append(row)

    spaced_duplicate_rows = []

    for normalized_id, grouped_rows in rows_by_normalized_id.items():
        has_spaced_version = any(
            has_spaced_dash(get_cell_value(row, document_id_index))
            for row in grouped_rows
        )

        has_unspaced_version = any(
            not has_spaced_dash(get_cell_value(row, document_id_index))
            for row in grouped_rows
        )

        if has_spaced_version and has_unspaced_version:
            spaced_duplicate_rows.extend(grouped_rows)

    spaced_duplicate_rows = sorted(
        spaced_duplicate_rows,
        key=lambda row: document_id_sort_key(get_cell_value(row, document_id_index)),
    )

    review_header = build_review_header(
        header=header,
        document_id_index=document_id_index,
        legacy_document_index=legacy_document_index,
    )

    spaced_dash_output_rows = [
        review_header,
        *[
            build_review_row(
                row=row,
                document_id_index=document_id_index,
                legacy_document_index=legacy_document_index,
                header_length=len(header),
            )
            for row in spaced_dash_rows
        ],
    ]

    spaced_duplicate_output_rows = [
        review_header,
        *[
            build_review_row(
                row=row,
                document_id_index=document_id_index,
                legacy_document_index=legacy_document_index,
                header_length=len(header),
            )
            for row in spaced_duplicate_rows
        ],
    ]

    workbook = Workbook()

    sorted_sheet = workbook.active
    sorted_sheet.title = "Sorted Document IDs"

    spaced_dash_sheet = workbook.create_sheet("Spaced Dashing")
    spaced_duplicate_sheet = workbook.create_sheet("Spaced Duplicates")

    write_rows_to_sheet(sorted_sheet, [header, *sorted_rows])
    write_rows_to_sheet(spaced_dash_sheet, spaced_dash_output_rows)
    write_rows_to_sheet(spaced_duplicate_sheet, spaced_duplicate_output_rows)

    style_worksheet(sorted_sheet)
    style_worksheet(spaced_dash_sheet)
    style_worksheet(spaced_duplicate_sheet)

    workbook.save(output_xlsx_path)

    print("Done.")
    print(f"Sorted rows written: {len(sorted_rows)}")
    print(f"Spaced dash rows written: {len(spaced_dash_rows)}")
    print(f"Spaced duplicate rows written: {len(spaced_duplicate_rows)}")
    print(f"Output written to: {output_xlsx_path}")


if __name__ == "__main__":
    main()