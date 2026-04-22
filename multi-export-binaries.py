#!/usr/bin/env python3
"""
DMS Multi-File Folder Finder (GUI) — Excel + PDF Reporting

This revision restores full event-log printing to the GUI output pane and
keeps the PDF fixes:
  - No zero-width characters inserted (prevents black blocks in PDFs)
  - Wrapping via ParagraphStyle(wordWrap="CJK", splitLongWords=1)
  - Table widths constrained to printable width
  - Footer shows "Page X of N"
  - Parent key padding rules: PAD5 for selected prefixes; preserve WI width
"""

from __future__ import annotations

import csv
import re
import threading
import queue
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# PDF (ReportLab)
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    LongTable,
)
from reportlab.pdfgen import canvas as rl_canvas


INCLUDE_ROLES_DEFAULT = {"binary", "rendition"}

# Prefixes normalized to 5-digit document number (middle chunk).
PAD5_PREFIXES = {"CSIL", "BOM", "COT", "OJT", "RPT", "SLIA", "SMP", "SSP"}


@dataclass(frozen=True)
class GranularHit:
    parent_key: str
    doc_folder: str
    version_folder: str
    folder_type: str
    folder_rel: str
    file_count: int


def _soft_wrap_text(s: str) -> str:
    """
    IMPORTANT:
    Do NOT insert any zero-width characters. Some PDF viewers render them as black blocks.
    Wrapping is handled by ReportLab ParagraphStyle(wordWrap="CJK", splitLongWords=1).
    """
    return "" if s is None else str(s)


def shorten_doc_folder(doc_folder: str) -> str:
    """
    Example folders:
      CSIL-0001-000001  -> CSIL-00001   (pad to 5 for CSIL)
      BOM-0006-000001   -> BOM-00006    (pad to 5 for BOM)
      WI-0008-000001    -> WI-0008      (DO NOT pad for WI)

    Rule:
      - If prefix (letters before first dash) is in PAD5_PREFIXES => zfill(5)
      - Otherwise preserve the middle chunk exactly as-is (no added zeros)
    """
    m = re.match(r"^([A-Za-z]+)-(\d+)-(\d+)$", doc_folder)
    if not m:
        return doc_folder

    prefix_letters = m.group(1)
    doc_no = m.group(2)
    prefix = f"{prefix_letters}-"

    if prefix_letters in PAD5_PREFIXES:
        doc_norm = doc_no.zfill(5)
    else:
        doc_norm = doc_no

    return f"{prefix}{doc_norm}"


def build_parent_key(doc_folder: str, version_folder: str) -> str:
    doc_short = shorten_doc_folder(doc_folder)
    try:
        v2 = f"{int(version_folder):02d}"
    except Exception:
        v2 = version_folder
    return f"{doc_short}/{v2}"


def direct_files_count(folder: Path) -> int:
    try:
        return sum(1 for p in folder.iterdir() if p.is_file())
    except FileNotFoundError:
        return -1
    except PermissionError:
        return -2


def rel_folder_str(folder: Path, root: Path) -> str:
    try:
        rel = folder.resolve().relative_to(root.resolve()).as_posix()
    except Exception:
        rel = str(folder.resolve())
    if not rel.endswith("/"):
        rel += "/"
    return rel


def coverage_check(granular_hits: list[GranularHit], summary_keys: list[str]) -> tuple[bool, list[str], list[str]]:
    granular_keys = sorted({h.parent_key for h in granular_hits})
    summary_set = set(summary_keys)
    granular_set = set(granular_keys)

    missing = sorted(granular_set - summary_set)
    extra = sorted(summary_set - granular_set)
    ok = (len(missing) == 0 and len(extra) == 0)
    return ok, missing, extra


def _autofit_columns(ws, min_width: int = 10, max_width: int = 80) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def write_excel_report(
    out_xlsx: Path,
    report_id: str,
    run_ts: datetime,
    raw_export_root: Path,
    csv_path: Path,
    output_dir: Path,
    effective_col: str,
    effective_value: str,
    role_col: str,
    input_path_col: str,
    include_roles: list[str],
    total_rows: int,
    kept_rows: int,
    skipped_status: int,
    skipped_role: int,
    skipped_missing_path: int,
    skipped_lastuploaded: int,
    skipped_wrong_parent: int,
    candidate_folders: int,
    granular_hits_sorted: list[GranularHit],
    unique_parent_keys: list[str],
    coverage_ok: bool,
    coverage_missing: list[str],
    coverage_extra: list[str],
    missing_folders: int,
    perm_folders: int,
) -> None:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_gran = wb.create_sheet("Granular")

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    r = 1
    ws_sum.cell(r, 1, "Multi-File Folder Finder Report").font = Font(bold=True, size=14)
    r += 2

    summary_kv = [
        ("Report ID", report_id),
        ("Run Timestamp (local)", run_ts.strftime("%Y-%m-%d %H:%M:%S")),
        ("Raw DMS Export Root", str(raw_export_root)),
        ("Input CSV", str(csv_path)),
        ("Output Folder", str(output_dir)),
        ("EFFECTIVE Column", effective_col),
        ("EFFECTIVE Value", effective_value),
        ("Role Column", role_col),
        ("Path Column", input_path_col),
        ("Include Roles", ", ".join(include_roles)),
        ("Total CSV rows read", total_rows),
        ("Rows kept after status+role filters", kept_rows),
        (f"Rows skipped (status != {effective_value})", skipped_status),
        (f"Rows skipped (role not in {include_roles})", skipped_role),
        (f"Rows skipped (missing {input_path_col})", skipped_missing_path),
        ("Rows skipped (path under lastuploaded)", skipped_lastuploaded),
        ("Rows skipped (parent not binaries/renditions)", skipped_wrong_parent),
        ("Candidate folders (unique)", candidate_folders),
        ("Granular hits printed", len(granular_hits_sorted)),
        ("Total unique multi-binaries found", len(unique_parent_keys)),
        ("Coverage check", "PASS" if coverage_ok else "FAIL"),
        ("Candidate folders missing on disk", missing_folders),
        ("Candidate folders permission denied", perm_folders),
    ]

    ws_sum.cell(r, 1, "Run Statistics").font = Font(bold=True, size=12)
    r += 1
    ws_sum.cell(r, 1, "Field").font = bold
    ws_sum.cell(r, 2, "Value").font = bold
    r += 1

    for k, v in summary_kv:
        ws_sum.cell(r, 1, k).alignment = wrap
        ws_sum.cell(r, 2, v).alignment = wrap
        r += 1

    if not coverage_ok:
        r += 1
        ws_sum.cell(r, 1, "Coverage Missing in Summary").font = bold
        r += 1
        for k in coverage_missing:
            ws_sum.cell(r, 1, k)
            r += 1
        r += 1
        ws_sum.cell(r, 1, "Coverage Extra in Summary").font = bold
        r += 1
        for k in coverage_extra:
            ws_sum.cell(r, 1, k)
            r += 1

    r += 2
    ws_sum.cell(r, 1, "Unique Parent Folders (Summary List)").font = Font(bold=True, size=12)
    r += 1
    ws_sum.cell(r, 1, "parent_key").font = bold
    r += 1
    for k in unique_parent_keys:
        ws_sum.cell(r, 1, k)
        r += 1

    ws_sum.freeze_panes = "A4"
    _autofit_columns(ws_sum)

    headers = ["parent_key", "folder_rel", "folder_type", "file_count", "doc_folder", "version_folder"]
    for c, h in enumerate(headers, start=1):
        ws_gran.cell(1, c, h).font = bold

    row = 2
    for hit in granular_hits_sorted:
        ws_gran.cell(row, 1, hit.parent_key)
        ws_gran.cell(row, 2, hit.folder_rel)
        ws_gran.cell(row, 3, hit.folder_type)
        ws_gran.cell(row, 4, hit.file_count)
        ws_gran.cell(row, 5, hit.doc_folder)
        ws_gran.cell(row, 6, hit.version_folder)
        row += 1

    ws_gran.freeze_panes = "A2"
    _autofit_columns(ws_gran)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


class NumberedCanvas(rl_canvas.Canvas):
    """Two-pass canvas so footer can render 'Page X of N'."""
    def __init__(self, *args, report_id: str = "", **kwargs):
        super().__init__(*args, **kwargs)
        self._report_id = report_id
        self._saved_page_states: list[dict] = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        page_count = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_header_footer(page_count)
            super().showPage()
        super().save()

    def _draw_header_footer(self, page_count: int) -> None:
        self.saveState()
        self.setFont("Helvetica-Bold", 9)
        self.drawString(0.75 * inch, 10.75 * inch, f"Multi-File Folder Finder Report — {self._report_id}")
        self.setFont("Helvetica", 9)
        self.drawString(0.75 * inch, 0.65 * inch, f"Report ID: {self._report_id}")
        self.drawRightString(7.75 * inch, 0.65 * inch, f"Page {self.getPageNumber()} of {page_count}")
        self.restoreState()


def write_pdf_report(
    out_pdf: Path,
    report_id: str,
    run_ts: datetime,
    raw_export_root: Path,
    csv_path: Path,
    output_dir: Path,
    effective_col: str,
    effective_value: str,
    role_col: str,
    input_path_col: str,
    include_roles: list[str],
    total_rows: int,
    kept_rows: int,
    skipped_status: int,
    skipped_role: int,
    skipped_missing_path: int,
    skipped_lastuploaded: int,
    skipped_wrong_parent: int,
    candidate_folders: int,
    granular_hits_sorted: list[GranularHit],
    unique_parent_keys: list[str],
    coverage_ok: bool,
    coverage_missing: list[str],
    coverage_extra: list[str],
    missing_folders: int,
    perm_folders: int,
) -> None:
    styles = getSampleStyleSheet()
    style_title = styles["Title"]
    style_h2 = styles["Heading2"]

    body9 = ParagraphStyle(
        "body9",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        wordWrap="CJK",
        splitLongWords=1,
    )
    body8 = ParagraphStyle(
        "body8",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
        wordWrap="CJK",
        splitLongWords=1,
    )

    printable_width = 7.0 * inch

    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(out_pdf),
        pagesize=LETTER,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=1.05 * inch,
        bottomMargin=0.95 * inch,
        title=f"Multi-File Folder Finder Report {report_id}",
    )

    story: list = []
    story.append(Paragraph("Multi-File Folder Finder Report", style_title))
    story.append(Paragraph(f"Report ID: <b>{report_id}</b>", body9))
    story.append(Paragraph(f"Run Timestamp (local): {run_ts.strftime('%Y-%m-%d %H:%M:%S')}", body9))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Run Statistics", style_h2))

    stats_rows = [
        ["Field", "Value"],
        ["Raw DMS Export Root", str(raw_export_root)],
        ["Input CSV", str(csv_path)],
        ["Output Folder", str(output_dir)],
        ["EFFECTIVE Column", effective_col],
        ["EFFECTIVE Value", effective_value],
        ["Role Column", role_col],
        ["Path Column", input_path_col],
        ["Include Roles", ", ".join(include_roles)],
        ["Total CSV rows read", str(total_rows)],
        ["Rows kept after status+role filters", str(kept_rows)],
        [f"Rows skipped (status != {effective_value})", str(skipped_status)],
        [f"Rows skipped (role not in {include_roles})", str(skipped_role)],
        [f"Rows skipped (missing {input_path_col})", str(skipped_missing_path)],
        ["Rows skipped (path under lastuploaded)", str(skipped_lastuploaded)],
        ["Rows skipped (parent not binaries/renditions)", str(skipped_wrong_parent)],
        ["Candidate folders (unique)", str(candidate_folders)],
        ["Granular hits printed", str(len(granular_hits_sorted))],
        ["Total unique multi-binaries found", str(len(unique_parent_keys))],
        ["Coverage check", "PASS" if coverage_ok else "FAIL"],
        ["Candidate folders missing on disk", str(missing_folders)],
        ["Candidate folders permission denied", str(perm_folders)],
    ]

    stats_tbl = []
    for i, row in enumerate(stats_rows):
        if i == 0:
            stats_tbl.append([Paragraph(str(row[0]), body9), Paragraph(str(row[1]), body9)])
        else:
            stats_tbl.append([Paragraph(_soft_wrap_text(row[0]), body9), Paragraph(_soft_wrap_text(row[1]), body9)])

    col1 = 2.2 * inch
    col2 = printable_width - col1
    t = Table(stats_tbl, colWidths=[col1, col2], repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Coverage Check Details", style_h2))
    story.append(Paragraph(
        "PASS: All unique IDs derived from the granular section are present in the summary list."
        if coverage_ok else
        "FAIL: The summary list and granular-derived IDs do not match. See missing/extra below.",
        body9
    ))

    if not coverage_ok:
        if coverage_missing:
            story.append(Spacer(1, 6))
            story.append(Paragraph("Missing in Summary (present in granular):", body9))
            miss_rows = [["parent_key"]] + [[k] for k in coverage_missing]
            miss_tbl = [[Paragraph(_soft_wrap_text(str(x)), body9) for x in row] for row in miss_rows]
            miss_t = LongTable(miss_tbl, colWidths=[printable_width], repeatRows=1)
            miss_t.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(miss_t)

        if coverage_extra:
            story.append(Spacer(1, 6))
            story.append(Paragraph("Extra in Summary (not present in granular):", body9))
            extra_rows = [["parent_key"]] + [[k] for k in coverage_extra]
            extra_tbl = [[Paragraph(_soft_wrap_text(str(x)), body9) for x in row] for row in extra_rows]
            extra_t = LongTable(extra_tbl, colWidths=[printable_width], repeatRows=1)
            extra_t.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(extra_t)

    story.append(PageBreak())

    story.append(Paragraph("Unique Parent Folders (Summary List)", style_h2))
    summary_rows = [["parent_key"]] + [[k] for k in unique_parent_keys]
    summary_tbl = [[Paragraph(_soft_wrap_text(str(x)), body9) for x in row] for row in summary_rows]
    t2 = LongTable(summary_tbl, colWidths=[printable_width], repeatRows=1)
    t2.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t2)

    story.append(PageBreak())

    story.append(Paragraph("Granular Folder Locations", style_h2))

    w_parent = 1.35 * inch
    w_type = 0.85 * inch
    w_count = 0.75 * inch
    w_rel = printable_width - (w_parent + w_type + w_count)

    gran_rows = [["parent_key", "folder_rel", "folder_type", "file_count"]]
    for hit in granular_hits_sorted:
        gran_rows.append([hit.parent_key, hit.folder_rel, hit.folder_type, str(hit.file_count)])

    gran_tbl = []
    for i, row in enumerate(gran_rows):
        if i == 0:
            gran_tbl.append([Paragraph(str(c), body9) for c in row])
        else:
            gran_tbl.append([
                Paragraph(_soft_wrap_text(str(row[0])), body8),
                Paragraph(_soft_wrap_text(str(row[1])), body8),
                Paragraph(_soft_wrap_text(str(row[2])), body8),
                Paragraph(_soft_wrap_text(str(row[3])), body8),
            ])

    t3 = LongTable(gran_tbl, colWidths=[w_parent, w_rel, w_type, w_count], repeatRows=1)
    t3.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t3)

    doc.build(story, canvasmaker=lambda *a, **k: NumberedCanvas(*a, report_id=report_id, **k))


def analyze(
    raw_export_root: Path,
    csv_path: Path,
    output_dir: Path,
    effective_col: str,
    effective_value: str,
    role_col: str,
    input_path_col: str,
    include_roles: set[str],
    logq: "queue.Queue[str]",
) -> None:
    def log(line: str = "") -> None:
        logq.put(line)

    run_ts = datetime.now()
    report_id = f"Multi_report_{run_ts.strftime('%d%m%Y_%H%M%S')}"

    raw_export_root = raw_export_root.expanduser().resolve()
    csv_path = csv_path.expanduser().resolve()
    output_dir = output_dir.expanduser().resolve()

    if not raw_export_root.exists() or not raw_export_root.is_dir():
        log(f"ERROR: Raw export root not found / not a directory: {raw_export_root}")
        return
    if not csv_path.exists():
        log(f"ERROR: CSV not found: {csv_path}")
        return
    if not output_dir.exists() or not output_dir.is_dir():
        log(f"ERROR: Output directory not found / not a directory: {output_dir}")
        return

    log("Starting analysis...")
    log("")
    log("Scanning CSV to build in-scope candidate folders ...")

    candidate_folders: set[Path] = set()
    total_rows = 0
    kept_rows = 0
    skipped_lastuploaded = 0
    skipped_wrong_parent = 0
    skipped_missing_path = 0
    skipped_status = 0
    skipped_role = 0

    include_roles_lc = {r.lower() for r in include_roles}
    eff_val_uc = effective_value.strip().upper()

    with csv_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []

        missing_cols = [c for c in (effective_col, role_col, input_path_col) if c not in cols]
        if missing_cols:
            log("ERROR: CSV missing required columns: " + ", ".join(missing_cols))
            log("Found columns include: " + ", ".join(cols))
            return

        for row in reader:
            total_rows += 1

            status = (row.get(effective_col) or "").strip().upper()
            if status != eff_val_uc:
                skipped_status += 1
                continue

            role = (row.get(role_col) or "").strip().lower()
            if role not in include_roles_lc:
                skipped_role += 1
                continue

            input_abs = (row.get(input_path_col) or "").strip()
            if not input_abs:
                skipped_missing_path += 1
                continue

            p = Path(input_abs)

            if any(part.lower() == "lastuploaded" for part in p.parts):
                skipped_lastuploaded += 1
                continue

            folder = p.parent
            if folder.name.lower() not in ("binaries", "renditions"):
                skipped_wrong_parent += 1
                continue

            candidate_folders.add(folder)
            kept_rows += 1

    log("---- CSV FILTER SUMMARY ----")
    log(f"Total CSV rows read                          : {total_rows}")
    log(f"Rows kept after status+role filters          : {kept_rows}")
    log(f"Rows skipped (status != {effective_value})          : {skipped_status}")
    log(f"Rows skipped (role not in {sorted(include_roles_lc)}) : {skipped_role}")
    log(f"Rows skipped (missing {input_path_col})            : {skipped_missing_path}")
    log(f"Rows skipped (path under lastuploaded)        : {skipped_lastuploaded}")
    log(f"Rows skipped (parent not binaries/renditions) : {skipped_wrong_parent}")
    log(f"Candidate folders (unique)                   : {len(candidate_folders)}")
    log("---------------------------")
    log("")

    if not candidate_folders:
        log("No candidate binaries/ or renditions/ folders were derived from the CSV filters.")
        return

    log("Scanning filesystem for multi-file folders (direct files only; no recursion) ...")
    log("")

    granular_hits: list[GranularHit] = []
    parentkey_to_anyhit: dict[str, bool] = {}
    missing_folders = 0
    perm_folders = 0

    for folder in sorted(candidate_folders):
        cnt = direct_files_count(folder)
        if cnt == -1:
            missing_folders += 1
            continue
        if cnt == -2:
            perm_folders += 1
            continue
        if cnt <= 1:
            continue

        version_folder = folder.parent.name if folder.parent else "?"
        doc_folder = folder.parent.parent.name if folder.parent and folder.parent.parent else "?"
        parent_key = build_parent_key(doc_folder, version_folder)

        granular_hits.append(GranularHit(
            parent_key=parent_key,
            doc_folder=doc_folder,
            version_folder=version_folder,
            folder_type=folder.name.lower(),
            folder_rel=rel_folder_str(folder, raw_export_root),
            file_count=cnt,
        ))
        parentkey_to_anyhit[parent_key] = True

    granular_hits_sorted = sorted(granular_hits, key=lambda h: (h.parent_key, h.folder_rel))
    unique_parent_keys = sorted(parentkey_to_anyhit.keys())
    ok, missing, extra = coverage_check(granular_hits, unique_parent_keys)

    log("---- COVERAGE CHECK ----")
    log(f"Granular hits printed                      : {len(granular_hits_sorted)}")
    log(f"Unique parent IDs derived from granular    : {len(set(h.parent_key for h in granular_hits))}")
    log(f"Unique parent IDs printed in summary       : {len(unique_parent_keys)}")
    log("Coverage check                             : " + ("PASS (missing=0, extra=0)" if ok else f"FAIL (missing={len(missing)}, extra={len(extra)})"))
    log("------------------------")
    log("")

    log(f"Total unique multi-binaries found: {len(unique_parent_keys)}")
    log("List of unique parent folders:")
    for k in unique_parent_keys:
        log(f"  {k}")

    log("")
    log("Granular folder locations:")
    for hit in granular_hits_sorted:
        log(f"{hit.folder_rel:<40} -> {hit.file_count} files")

    log("")
    log("Filesystem anomalies (informational):")
    log(f"  Candidate folders missing on disk     : {missing_folders}")
    log(f"  Candidate folders permission denied   : {perm_folders}")

    out_xlsx = output_dir / f"{report_id}.xlsx"
    out_pdf = output_dir / f"{report_id}.pdf"

    try:
        log("")
        log(f"Writing Excel report: {out_xlsx}")
        write_excel_report(
            out_xlsx=out_xlsx,
            report_id=report_id,
            run_ts=run_ts,
            raw_export_root=raw_export_root,
            csv_path=csv_path,
            output_dir=output_dir,
            effective_col=effective_col,
            effective_value=effective_value,
            role_col=role_col,
            input_path_col=input_path_col,
            include_roles=sorted(include_roles_lc),
            total_rows=total_rows,
            kept_rows=kept_rows,
            skipped_status=skipped_status,
            skipped_role=skipped_role,
            skipped_missing_path=skipped_missing_path,
            skipped_lastuploaded=skipped_lastuploaded,
            skipped_wrong_parent=skipped_wrong_parent,
            candidate_folders=len(candidate_folders),
            granular_hits_sorted=granular_hits_sorted,
            unique_parent_keys=unique_parent_keys,
            coverage_ok=ok,
            coverage_missing=missing,
            coverage_extra=extra,
            missing_folders=missing_folders,
            perm_folders=perm_folders,
        )

        log(f"Writing PDF report:   {out_pdf}")
        write_pdf_report(
            out_pdf=out_pdf,
            report_id=report_id,
            run_ts=run_ts,
            raw_export_root=raw_export_root,
            csv_path=csv_path,
            output_dir=output_dir,
            effective_col=effective_col,
            effective_value=effective_value,
            role_col=role_col,
            input_path_col=input_path_col,
            include_roles=sorted(include_roles_lc),
            total_rows=total_rows,
            kept_rows=kept_rows,
            skipped_status=skipped_status,
            skipped_role=skipped_role,
            skipped_missing_path=skipped_missing_path,
            skipped_lastuploaded=skipped_lastuploaded,
            skipped_wrong_parent=skipped_wrong_parent,
            candidate_folders=len(candidate_folders),
            granular_hits_sorted=granular_hits_sorted,
            unique_parent_keys=unique_parent_keys,
            coverage_ok=ok,
            coverage_missing=missing,
            coverage_extra=extra,
            missing_folders=missing_folders,
            perm_folders=perm_folders,
        )

        log("")
        log(f"Saved Excel report: {out_xlsx}")
        log(f"Saved PDF report:   {out_pdf}")

    except Exception as e:
        log("")
        log("ERROR: Failed generating Excel/PDF reports.")
        log(f"       {type(e).__name__}: {e}")


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("DMS Multi-File Folder Finder")
        self.geometry("1180x800")

        self.logq: "queue.Queue[str]" = queue.Queue()
        self.worker: threading.Thread | None = None

        self.var_root = tk.StringVar()
        self.var_csv = tk.StringVar()
        self.var_out_dir = tk.StringVar()

        self.var_effective_col = tk.StringVar(value="status_bucket")
        self.var_effective_val = tk.StringVar(value="EFFECTIVE")
        self.var_role_col = tk.StringVar(value="file_role")
        self.var_input_col = tk.StringVar(value="input_abs")
        self.var_roles = tk.StringVar(value="binary,rendition")

        self._build_ui()
        self.after(100, self._drain_logq)

    def _build_ui(self) -> None:
        pad = {"padx": 10, "pady": 7}

        frm = ttk.Frame(self)
        frm.pack(fill="x", **pad)

        ttk.Label(frm, text="Raw DMS Export Root:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.var_root, width=120).grid(row=0, column=1, sticky="we")
        ttk.Button(frm, text="Browse...", command=self._browse_dir).grid(row=0, column=2, sticky="e")

        ttk.Label(frm, text="Neurotic_DMS_org_Manifest CSV:").grid(row=1, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.var_csv, width=120).grid(row=1, column=1, sticky="we")
        ttk.Button(frm, text="Browse...", command=self._browse_csv).grid(row=1, column=2, sticky="e")

        ttk.Label(frm, text="Output folder (auto-generates .xlsx and .pdf):").grid(row=2, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.var_out_dir, width=120).grid(row=2, column=1, sticky="we")
        ttk.Button(frm, text="Browse...", command=self._browse_out_dir).grid(row=2, column=2, sticky="e")

        frm.columnconfigure(1, weight=1)

        filters = ttk.LabelFrame(self, text="CSV Filters")
        filters.pack(fill="x", **pad)

        ttk.Label(filters, text="EFFECTIVE column:").grid(row=0, column=0, sticky="w", padx=10, pady=4)
        ttk.Entry(filters, textvariable=self.var_effective_col, width=28).grid(row=0, column=1, sticky="w", padx=10, pady=4)

        ttk.Label(filters, text="EFFECTIVE value:").grid(row=0, column=2, sticky="w", padx=10, pady=4)
        ttk.Entry(filters, textvariable=self.var_effective_val, width=18).grid(row=0, column=3, sticky="w", padx=10, pady=4)

        ttk.Label(filters, text="Role column:").grid(row=1, column=0, sticky="w", padx=10, pady=4)
        ttk.Entry(filters, textvariable=self.var_role_col, width=28).grid(row=1, column=1, sticky="w", padx=10, pady=4)

        ttk.Label(filters, text="Path column:").grid(row=1, column=2, sticky="w", padx=10, pady=4)
        ttk.Entry(filters, textvariable=self.var_input_col, width=18).grid(row=1, column=3, sticky="w", padx=10, pady=4)

        ttk.Label(filters, text="Include roles (comma-separated):").grid(row=2, column=0, sticky="w", padx=10, pady=4)
        ttk.Entry(filters, textvariable=self.var_roles, width=60).grid(row=2, column=1, columnspan=3, sticky="w", padx=10, pady=4)

        btns = ttk.Frame(self)
        btns.pack(fill="x", **pad)

        self.btn_run = ttk.Button(btns, text="Run Analysis", command=self._run)
        self.btn_run.pack(side="left")

        ttk.Button(btns, text="Clear Output", command=self._clear).pack(side="left", padx=10)

        log_frame = ttk.LabelFrame(self, text="Output")
        log_frame.pack(fill="both", expand=True, **pad)

        self.txt = tk.Text(log_frame, wrap="none")
        self.txt.pack(side="left", fill="both", expand=True)

        yscroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.txt.yview)
        yscroll.pack(side="right", fill="y")
        self.txt.configure(yscrollcommand=yscroll.set)

        xscroll = ttk.Scrollbar(self, orient="horizontal", command=self.txt.xview)
        xscroll.pack(fill="x", padx=10)
        self.txt.configure(xscrollcommand=xscroll.set)

    def _browse_dir(self) -> None:
        p = filedialog.askdirectory(title="Select Raw DMS Export Root")
        if p:
            self.var_root.set(p)

    def _browse_csv(self) -> None:
        p = filedialog.askopenfilename(
            title="Select Neurotic_DMS_org_Manifest CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if p:
            self.var_csv.set(p)

    def _browse_out_dir(self) -> None:
        p = filedialog.askdirectory(title="Select Output Folder (reports auto-named)")
        if p:
            self.var_out_dir.set(p)

    def _clear(self) -> None:
        self.txt.delete("1.0", "end")

    def _append(self, line: str) -> None:
        self.txt.insert("end", line + "\n")
        self.txt.see("end")

    def _drain_logq(self) -> None:
        try:
            while True:
                line = self.logq.get_nowait()
                self._append(line)
        except queue.Empty:
            pass
        self.after(100, self._drain_logq)

    def _run(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Running", "Analysis is already running.")
            return

        root = self.var_root.get().strip()
        csvp = self.var_csv.get().strip()
        outd = self.var_out_dir.get().strip()

        if not root or not csvp or not outd:
            messagebox.showerror("Missing Input", "Please provide Raw DMS Export Root, Neurotic_DMS_org_Manifest CSV, and Output Folder.")
            return

        effective_col = self.var_effective_col.get().strip()
        effective_val = self.var_effective_val.get().strip()
        role_col = self.var_role_col.get().strip()
        input_col = self.var_input_col.get().strip()

        roles_raw = self.var_roles.get().strip()
        roles = {r.strip().lower() for r in roles_raw.split(",") if r.strip()}
        if not roles:
            roles = set(INCLUDE_ROLES_DEFAULT)

        self.btn_run.configure(state="disabled")
        self._append("Starting analysis...\n")

        def worker_fn() -> None:
            try:
                analyze(
                    raw_export_root=Path(root),
                    csv_path=Path(csvp),
                    output_dir=Path(outd),
                    effective_col=effective_col,
                    effective_value=effective_val,
                    role_col=role_col,
                    input_path_col=input_col,
                    include_roles=roles,
                    logq=self.logq,
                )
            finally:
                self.after(0, lambda: self.btn_run.configure(state="normal"))

        self.worker = threading.Thread(target=worker_fn, daemon=True)
        self.worker.start()


if __name__ == "__main__":
    App().mainloop()
