"""
Remove the standard Legacy Migrated Record Disclaimer from every text cell
in every worksheet of an Excel workbook.

What this script does:
- Opens an Excel workbook
- Scans every text cell in every worksheet
- Removes the exact disclaimer text wherever it appears
- Cleans up extra blank space left behind
- Saves a new cleaned workbook

Edit these two lines first if needed:
- INPUT_CSV
- OUTPUT_CSV
"""

from pathlib import Path
from openpyxl import load_workbook


# =========================
# EDIT THESE PATHS
# =========================
INPUT_CSV = r"C:\Users\e.muraj\OneDrive - Neurotech USA, Inc\production26-csv\legacy QMS summaries.xlsx"
OUTPUT_CSV = r"C:\Users\e.muraj\OneDrive - Neurotech USA, Inc\production26-csv\legacy QMS summaries_clean.xlsx"


DISCLAIMER = (
    "Legacy Migrated Record Disclaimer: This file represents a reconstructed digital record "
    "derived from a TrackWise® Quality Management System export. TrackWise stores GMP records "
    "across many inter-related .csv files and associated binary attachments, with information "
    "distributed across tables using cryptic system-generated identifiers. The contents of this "
    "reconstructed record were generated through a deterministic and fully auditable algorithmic "
    "process that parsed relevant TrackWise source tables, resolved foreign-key relationships, "
    "reconstructed the hierarchical record structure, retrieved associated original file "
    "attachments (“blobs”), produced a structured JSON representation, and generated a DOCX "
    "summary and compiled the reconstructed summary plus associated files into a single, "
    "paginated final PDF suitable for inspection and long-term archival. The “Additional Files” "
    "provided alongside this record are the standalone source components used in compilation, "
    "including the reconstructed JSON, the DOCX summary, and the original attachments; these "
    "items are provided individually for traceability and—after conversion where applicable—also "
    "comprise the merged final PDF. No information was inferred, altered, or omitted. CR-00147 "
    "is the governing Change Control that authorizes and defines the migration/reconstruction of "
    "legacy TrackWise records into the new QMS record format and archival package."
)


def clean_cell(value):
    """Remove disclaimer from a single cell and tidy leftover whitespace."""
    if value is None:
        return value

    if not isinstance(value, str):
        return value

    text = value

    if DISCLAIMER in text:
        text = text.replace(DISCLAIMER, "")

        # Clean up leftover spacing/newlines from removal
        text = text.replace("\r\n", "\n")
        text = text.replace("\r", "\n")

        # Remove triple+ blank lines repeatedly
        while "\n\n\n" in text:
            text = text.replace("\n\n\n", "\n\n")

        # Trim whitespace at ends
        text = text.strip()

    return text


def main():
    input_path = Path(INPUT_CSV)
    output_path = Path(OUTPUT_CSV)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    wb = load_workbook(input_path)
    changed_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                original_value = cell.value
                new_value = clean_cell(original_value)

                if new_value != original_value:
                    cell.value = new_value
                    changed_cells += 1

    wb.save(output_path)

    print("Done.")
    print(f"Original file: {input_path}")
    print(f"Cleaned file : {output_path}")
    print(f"Cells changed: {changed_cells}")


if __name__ == "__main__":
    main()