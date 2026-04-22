#!/usr/bin/env python3
# training_history_migration_gui.py
r"""
Training History Migration Builder (Tkinter)

WHAT IT DOES
- User selects ONE input parent folder containing Excel files (.xls and/or .xlsx)
- Tool discovers all matching files in that folder (non-recursive) and processes them
- User selects an output folder
- Option: one output file per input file OR one consolidated output file for all inputs
- Produces an inspection-ready evidence package under a deterministic run folder:

  RUN_NAME = TRAINING_HISTORY_MIGRATION_MMDDYYYY_HHMMSS
  RUN_DIR  = <OutputFolder>\<RUN_NAME>\

OUTPUTS
- Output XLSX(s): headers row 1, data starts row 2 (clean sheet)
- PDF run report:
  * fixed-width "run log" style (Courier) like your other widgets
  * footer ONLY on every page: "<RUN_NAME>" (left) and "Page X of N" (right)
  * includes provenance, per-file verification (hashed vs confirmed), status distributions, fail summary, exceptions
- CSV audit logs:
  * <RUN_NAME>_run_manifest.csv
  * <RUN_NAME>_row_evidence.csv      (Option A: combined evidence chain input hash -> output hash)
  * <RUN_NAME>_exceptions.csv

INPUT SUPPORT
- .xlsx read via openpyxl
- .xls auto-converted to .xlsx using Microsoft Excel COM automation (Windows + Excel required)
  - Requires: pip install pywin32

Dependencies
  pip install openpyxl reportlab pywin32
"""

from __future__ import annotations

import csv
import getpass
import hashlib
import os
import platform
import queue
import re
import socket
import sys
import threading
import time
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas as rl_canvas


# =========================
# Constants / Schema
# =========================

RUN_PREFIX = "TRAINING_HISTORY_MIGRATION"
HEADER_SCAN_MAX_ROWS = 50
LOCAL_TZ_LABEL = "America/New_York"  # label only (system time used)

INPUT_REQUIRED_COLUMNS = [
    "Training Number",
    "Training Name",
    "Version",
    "Trainee Name",
    "Completed Date/Time",
    "Training Status",
]

OUTPUT_HEADERS = [
    "Training course name",
    "Revision",
    "Training course (macro set)",
    "Employee (Last name, First name)",
    "Trainer (Last name, First name)",
    "training completed on MM/DD/YYYY",
    "Pass/Fail",
    "Training record status",
    "Training record migration status",
    "Generate training record",
]

# Approved preface block (verbatim intent)
PDF_PREFACE = (
    "Purpose — This report documents an automated transformation of exported training-history spreadsheets into the "
    "target Training History Migration template. The input reports are simple training-history exports printed from "
    "TrackWise (i.e., standard history listings) and do not require any user modification, reformatting, or "
    "preprocessing prior to selection in this tool. The tool reads all Excel files in the selected input folder "
    "(.xls files are first converted to .xlsx via Microsoft Excel automation), detects the header row and required "
    "columns, and then processes each eligible data row using a deterministic mapping profile.\n\n"
    "Logic — For each eligible input row (non-blank across required fields), the tool computes an input-row hash to "
    "catalog the row as in-scope for processing. It then generates one output record with the following rules: "
    "Training course name is copied from Training Name; Revision is copied from Version with leading zeroes removed "
    "when numeric (e.g., “01”→“1”); Training course (macro set), Trainer, Training record migration status, and "
    "Generate training record are left blank; Employee (Last name, First name) is derived from Trainee Name by "
    "converting “First Last” to “Last, First” (if a comma already exists, the name is left unchanged; if the value is "
    "a single token it is left unchanged and logged as a warning); training completed on MM/DD/YYYY is derived from "
    "Completed Date/Time by stripping the time portion (splitting at the first comma and keeping the date token); "
    "Pass/Fail is set to “Pass” if Training Status contains the word “Completed” (case-insensitive substring), "
    "otherwise “Fail”; Training record status is set to “Passed” under the same condition, otherwise “Failed”. "
    "The tool also computes an output-row hash from the 10 output template fields and records both hashes in the "
    "row-level evidence file.\n\n"
    "SHA-256 hashing — SHA-256 is a one-way cryptographic hash function that converts a set of values into a fixed-length "
    "64-hex-character fingerprint. In this run, SHA-256 is used in three ways: (1) File hashes are computed for each input "
    "and each generated output/log file to provide a tamper-evident identifier for the exact artifacts used and produced; "
    "(2) an input-row hash is computed from the source row identity (file/sheet/row) and raw required input values to "
    "uniquely fingerprint each eligible row that should be processed; and (3) an output-row hash is computed from the final "
    "10 output fields written to the migration template. Because any change to the underlying values results in a different "
    "hash, these hashes allow objective verification that the correct inputs were used, that each eligible row was handled, "
    "and that the recorded evidence corresponds to the produced outputs.\n\n"
    "Verification — After processing, the tool reconciles row handling by confirming that every cataloged input-row hash is "
    "present in the processed set (i.e., “eligible rows hashed” equals “hashes confirmed processed” plus any intentionally "
    "skipped blank rows). It also records the associated output-row hash for each processed row, providing an explicit chain "
    "from input-row fingerprint to output-row fingerprint. Any mismatch is flagged as NOT RECONCILED and recorded in the "
    "exception log. This PDF is system-generated output produced at run time and summarizes the automated processing, "
    "reconciliation results, and output locations below."
)


# =========================
# Utility: versions, hashing, transforms
# =========================

def _safe_pkg_version(name: str) -> str:
    try:
        import importlib.metadata as importlib_metadata  # type: ignore
        return importlib_metadata.version(name)
    except Exception:
        return "unknown"


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            b = f.read(chunk_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def safe_filename(s: str, max_len: int = 90) -> str:
    s = re.sub(r"[^\w\-. ]+", "_", s).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:max_len] if len(s) > max_len else s


def strip_leading_zeroes(version: str) -> str:
    if version is None:
        return ""
    v = str(version).strip()
    if v == "":
        return ""
    if re.fullmatch(r"\d+", v):
        return str(int(v))
    return v


def name_first_last_to_last_first(name: str) -> Tuple[str, bool, str]:
    if name is None:
        return ("", False, "blank")
    raw = str(name).strip()
    if raw == "":
        return ("", False, "blank")
    if "," in raw:
        return (raw, False, "already_comma")
    parts = [p for p in raw.split() if p]
    if len(parts) == 1:
        return (raw, False, "single_token")
    last = parts[-1]
    first = " ".join(parts[:-1])
    return (f"{last}, {first}", True, "converted")


def strip_date_only(dt_text: str) -> Tuple[str, bool, str]:
    if dt_text is None:
        return ("", False, "blank")
    raw = str(dt_text).strip()
    if raw == "":
        return ("", False, "blank")
    if "," in raw:
        left = raw.split(",", 1)[0].strip()
        return (left, True, "comma_split")
    return (raw, False, "no_comma")


def contains_completed(status: str) -> bool:
    if status is None:
        return False
    return "completed" in str(status).lower()


def output_row_hash(values: List[str]) -> str:
    """SHA-256 of the 10 output template fields (written row payload)."""
    h = hashlib.sha256()
    joined = "\u241f".join([v if v is not None else "" for v in values])
    h.update(joined.encode("utf-8", errors="replace"))
    return h.hexdigest()


def input_row_hash(
    source_file: str,
    source_sheet: str,
    source_row: int,
    training_number: str,
    training_name: str,
    version_raw: str,
    trainee_raw: str,
    completed_raw: str,
    status_raw: str,
) -> str:
    """SHA-256 fingerprint of an eligible input row (includes row identity + required values)."""
    h = hashlib.sha256()
    payload = "\u241f".join([
        str(source_file or ""),
        str(source_sheet or ""),
        str(source_row),
        str(training_number or ""),
        str(training_name or ""),
        str(version_raw or ""),
        str(trainee_raw or ""),
        str(completed_raw or ""),
        str(status_raw or ""),
    ])
    h.update(payload.encode("utf-8", errors="replace"))
    return h.hexdigest()


# =========================
# XLS -> XLSX conversion (Excel COM)
# =========================

def convert_xls_to_xlsx_via_excel(src_xls: Path, dst_xlsx: Path) -> None:
    """
    Convert legacy .xls to .xlsx using Microsoft Excel COM automation (Windows only).
    Requires Excel installed and pywin32 installed.

      pip install pywin32
    """
    try:
        import win32com.client  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "Cannot convert .xls because pywin32 is not available.\n"
            "Install it with: pip install pywin32\n"
            f"Original error: {e}"
        )

    if dst_xlsx.exists():
        dst_xlsx.unlink()

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(src_xls))
        wb.SaveAs(str(dst_xlsx), FileFormat=51)  # 51 = .xlsx
    finally:
        if wb is not None:
            wb.Close(False)
        excel.Quit()


# =========================
# Audit dataclasses
# =========================

@dataclass
class ManifestRow:
    run_name: str
    role: str  # INPUT/OUTPUT/LOG/REPORT
    path: str
    size_bytes: int
    modified_local: str
    sha256: str


@dataclass
class ExceptionRow:
    run_name: str
    severity: str  # INFO/WARN/ERROR
    code: str
    message: str
    source_file: str
    source_sheet: str
    source_row: str
    column: str
    raw_value: str
    action: str


@dataclass
class RowEvidence:
    run_name: str
    source_file: str
    source_sheet: str
    source_row: int

    # Verification chain
    input_row_hash: str
    processed_flag: str  # Y/N

    output_file: str
    output_sheet: str
    output_row: int
    output_row_fingerprint_sha256: str

    # Raw / normalized fields
    training_number: str
    training_name: str
    version_raw: str
    revision_norm: str
    trainee_raw: str
    employee_norm: str
    completed_raw: str
    completed_date: str
    status_raw: str
    pass_fail: str
    record_status: str

    # Flags
    rev_zero_stripped: str
    name_reformatted: str
    date_stripped: str


@dataclass
class FileRunStats:
    source_file: str
    source_path: str
    source_type: str  # xls/xlsx
    converted_path: str  # "" if none
    sheet_used: str
    header_row: int
    data_start_row: int

    rows_scanned_total: int
    rows_blank_skipped: int
    eligible_rows_hashed: int
    hashes_confirmed_processed: int
    unaccounted_rows: int

    output_rows_written: int
    pass_count: int
    fail_count: int

    distinct_trainees: int
    distinct_courses: int

    date_min: str
    date_max: str

    reconciliation_status: str  # OK / NOT_RECONCILED


# =========================
# Fixed-width PDF builder ("run log style")
# =========================

class FixedWidthPDF:
    def __init__(
        self,
        pdf_path: Path,
        run_name: str,
        pagesize=letter,
        margin_left=0.75 * inch,
        margin_right=0.75 * inch,
        margin_top=0.75 * inch,
        margin_bottom=0.9 * inch,
        font_name="Courier",
        font_size=9,
        leading=11,
    ):
        self.pdf_path = pdf_path
        self.run_name = run_name
        self.pagesize = pagesize
        self.ml = margin_left
        self.mr = margin_right
        self.mt = margin_top
        self.mb = margin_bottom
        self.font_name = font_name
        self.font_size = font_size
        self.leading = leading

    def _compute_max_chars(self) -> int:
        c = rl_canvas.Canvas(str(self.pdf_path), pagesize=self.pagesize)
        c.setFont(self.font_name, self.font_size)
        usable_w = self.pagesize[0] - self.ml - self.mr
        # Courier is fixed width; estimate using "M"
        char_w = c.stringWidth("M", self.font_name, self.font_size)
        c.save()
        if char_w <= 0:
            return 90
        return max(60, int(usable_w // char_w))

    def _compute_lines_per_page(self) -> int:
        usable_h = self.pagesize[1] - self.mt - self.mb
        return max(20, int(usable_h // self.leading))

    @staticmethod
    def wrap_text(text: str, width: int, indent: int = 0) -> List[str]:
        """
        Simple hard wrap on whitespace; if a token exceeds width, it will be split.
        """
        if text is None:
            text = ""
        s = str(text).replace("\r\n", "\n").replace("\r", "\n")
        lines_out: List[str] = []
        for para in s.split("\n"):
            if para.strip() == "":
                lines_out.append("")
                continue
            words = para.split(" ")
            cur = " " * indent
            for w in words:
                if w == "":
                    continue
                if len(cur.rstrip()) == 0:
                    candidate = (" " * indent) + w
                else:
                    candidate = cur + " " + w
                if len(candidate) <= width:
                    cur = candidate
                else:
                    # flush current
                    if cur.strip():
                        lines_out.append(cur)
                    else:
                        # current is empty, must split word
                        pass
                    # split long word
                    while len(w) > (width - indent):
                        lines_out.append((" " * indent) + w[: (width - indent)])
                        w = w[(width - indent):]
                    cur = (" " * indent) + w
            if cur.strip() != "" or cur == (" " * indent):
                lines_out.append(cur.rstrip())
        return lines_out

    @staticmethod
    def make_table(
        headers: List[str],
        rows: List[List[str]],
        col_widths: List[int],
        sep_char: str = "-",
    ) -> List[str]:
        """
        Fixed-width table with wrapping inside cells.
        Assumes sum(col_widths) + separators fits max_chars.
        """
        def fmt_row(cols: List[str]) -> List[str]:
            wrapped_cols = []
            for i, cell in enumerate(cols):
                cell = "" if cell is None else str(cell)
                wrapped_cols.append(FixedWidthPDF.wrap_text(cell, col_widths[i], indent=0))
            max_lines = max(len(w) for w in wrapped_cols) if wrapped_cols else 1
            out_lines = []
            for li in range(max_lines):
                parts = []
                for ci in range(len(col_widths)):
                    part = wrapped_cols[ci][li] if li < len(wrapped_cols[ci]) else ""
                    parts.append(part.ljust(col_widths[ci])[:col_widths[ci]])
                out_lines.append(" | ".join(parts).rstrip())
            return out_lines

        # Header
        lines = []
        lines.extend(fmt_row(headers))
        total_w = sum(col_widths) + 3 * (len(col_widths) - 1)
        lines.append(sep_char * total_w)
        for r in rows:
            lines.extend(fmt_row(r))
        return lines

    def write(self, lines: List[str]):
        max_chars = self._compute_max_chars()
        lpp = self._compute_lines_per_page()

        # Wrap any overlong lines defensively
        expanded: List[str] = []
        for ln in lines:
            if ln is None:
                expanded.append("")
                continue
            s = str(ln)
            if len(s) <= max_chars:
                expanded.append(s)
            else:
                expanded.extend(self.wrap_text(s, max_chars, indent=0))

        pages: List[List[str]] = []
        cur: List[str] = []
        for ln in expanded:
            cur.append(ln)
            if len(cur) >= lpp:
                pages.append(cur)
                cur = []
        if cur:
            pages.append(cur)

        total_pages = max(1, len(pages))

        c = rl_canvas.Canvas(str(self.pdf_path), pagesize=self.pagesize)
        c.setFont(self.font_name, self.font_size)

        for page_idx, page_lines in enumerate(pages, start=1):
            # draw body
            x0 = self.ml
            y0 = self.pagesize[1] - self.mt
            y = y0
            for ln in page_lines:
                c.drawString(x0, y, ln)
                y -= self.leading

            # footer only (requirement)
            footer_y = 0.5 * inch
            c.setFont(self.font_name, 9)
            c.drawString(self.ml, footer_y, self.run_name)
            page_text = f"Page {page_idx} of {total_pages}"
            w = c.stringWidth(page_text, self.font_name, 9)
            c.drawString(self.pagesize[0] - self.mr - w, footer_y, page_text)

            c.showPage()
            c.setFont(self.font_name, self.font_size)

        c.save()


# =========================
# Transformer
# =========================

class TrainingTransformer:
    def __init__(
        self,
        run_name: str,
        run_dir: Path,
        log_q: queue.Queue,
        cancel_evt: threading.Event,
        pause_evt: threading.Event,
        input_dir: Optional[str] = None,
        output_dir: Optional[str] = None,
        mode: Optional[str] = None,
        tool_path: Optional[str] = None,
    ):
        self.run_name = run_name
        self.run_dir = run_dir
        self.log_q = log_q
        self.cancel_evt = cancel_evt
        self.pause_evt = pause_evt

        # Run context (provenance)
        self.run_started_ts = time.time()
        self.run_started_local = datetime.now().strftime(f"%m/%d/%Y %I:%M:%S %p {LOCAL_TZ_LABEL}")
        self.run_started_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        self.run_finished_local = ""
        self.run_finished_utc = ""
        self.run_duration_seconds = 0

        self.context = {
            "run_name": run_name,
            "input_dir": input_dir or "",
            "output_dir": output_dir or "",
            "mode": mode or "",
            "user": getpass.getuser(),
            "hostname": socket.gethostname(),
            "os": platform.platform(),
            "python": sys.version.replace("\n", " "),
            "python_executable": sys.executable,
            "cwd": str(Path.cwd()),
            "tool_path": tool_path or "",
            "openpyxl_version": _safe_pkg_version("openpyxl"),
            "reportlab_version": _safe_pkg_version("reportlab"),
            "pywin32_version": _safe_pkg_version("pywin32"),
            "started_local": self.run_started_local,
            "started_utc": self.run_started_utc,
        }

        # Logs / evidence
        self.exceptions: List[ExceptionRow] = []
        self.manifest: List[ManifestRow] = []
        self.file_stats: List[FileRunStats] = []

        # Evidence (Option A)
        self._evidence_by_hash: Dict[str, RowEvidence] = {}
        self._evidence_order: List[str] = []  # stable ordering of input_row_hash values

        # Verification sets
        self._eligible_hashes_by_key: Dict[Tuple[str, str], set] = {}   # (source_file, sheet) -> eligible input_row_hash set
        self._processed_hashes_by_key: Dict[Tuple[str, str], set] = {}  # (source_file, sheet) -> processed input_row_hash set

        # High-level aggregate stats
        self.stats = {
            "rows_blank_skipped_total": 0,
            "eligible_rows_hashed_total": 0,
            "hashes_confirmed_processed_total": 0,
            "output_rows_written_total": 0,
            "pass_total": 0,
            "fail_total": 0,
            "distinct_trainees": set(),
            "distinct_courses": set(),
            "date_min": None,
            "date_max": None,
            "rev_zero_stripped_count": 0,
            "name_converted_count": 0,
            "name_already_comma_count": 0,
            "name_single_token_count": 0,
            "date_comma_stripped_count": 0,
            "date_no_comma_count": 0,
            "date_parse_failed_count": 0,
            "status_dist": {},  # status_raw -> count
        }

    def _log(self, msg: str):
        self.log_q.put(("log", msg))

    def _check_pause_cancel(self):
        if self.cancel_evt.is_set():
            raise RuntimeError("Cancelled by user.")
        while self.pause_evt.is_set():
            time.sleep(0.1)
            if self.cancel_evt.is_set():
                raise RuntimeError("Cancelled by user.")

    def _add_exception(
        self,
        severity: str,
        code: str,
        message: str,
        source_file: str,
        source_sheet: str,
        source_row: str,
        column: str,
        raw_value: str,
        action: str,
    ):
        self.exceptions.append(ExceptionRow(
            run_name=self.run_name,
            severity=severity,
            code=code,
            message=message,
            source_file=source_file,
            source_sheet=source_sheet,
            source_row=str(source_row),
            column=column,
            raw_value=(str(raw_value)[:240] if raw_value is not None else ""),
            action=action,
        ))

    def _manifest_add_file(self, role: str, path: Path):
        st = path.stat()
        self.manifest.append(ManifestRow(
            run_name=self.run_name,
            role=role,
            path=str(path),
            size_bytes=int(st.st_size),
            modified_local=datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            sha256=sha256_file(path),
        ))

    def _parse_date_for_range(self, date_str: str) -> Optional[datetime]:
        if not date_str:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None

    def _normalize_input_to_xlsx(self, src_path: Path) -> Tuple[Path, str, str]:
        """Returns: (xlsx_path_to_read, display_name_original, converted_path_or_empty)."""
        suffix = src_path.suffix.lower()
        if suffix == ".xlsx":
            return src_path, src_path.name, ""
        if suffix == ".xls":
            conv_dir = self.run_dir / "_converted_inputs"
            conv_dir.mkdir(parents=True, exist_ok=True)
            dst = conv_dir / f"{src_path.stem}.xlsx"
            self._log(f"[convert] Converting .xls → .xlsx: {src_path.name} -> {dst.name}")
            convert_xls_to_xlsx_via_excel(src_path, dst)
            self._log(f"[convert] Conversion complete: {dst.name}")
            # record derived artifact in manifest as LOG
            self.manifest.append(ManifestRow(
                run_name=self.run_name,
                role="LOG",
                path=str(dst),
                size_bytes=int(dst.stat().st_size),
                modified_local=datetime.fromtimestamp(dst.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                sha256=sha256_file(dst),
            ))
            return dst, src_path.name, str(dst)
        raise ValueError(f"Unsupported input type: {src_path.name} (expected .xls or .xlsx)")

    def _write_output_xlsx(self, out_path: Path, rows: List[List[str]], sheet_name: str = "Training Migration"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(OUTPUT_HEADERS)  # row 1
        for r in rows:
            ws.append(r)           # row 2+
        ws.freeze_panes = "A2"
        last_col = chr(ord("A") + len(OUTPUT_HEADERS) - 1)
        ws.auto_filter.ref = f"A1:{last_col}{max(1, len(rows) + 1)}"
        wb.save(out_path)

    def _write_csv(self, path: Path, fieldnames: List[str], rows: List[dict]):
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for r in rows:
                w.writerow(r)

    def _detect_header_in_workbook(self, wb, display_name: str) -> Tuple[Worksheet, Dict[str, int], str, int]:
        required_norm = {c.strip().casefold() for c in INPUT_REQUIRED_COLUMNS}

        for ws in wb.worksheets:
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=HEADER_SCAN_MAX_ROWS, values_only=True), start=1):
                if row is None:
                    continue
                norm_to_idx: Dict[str, int] = {}
                for c_idx, val in enumerate(row):
                    if val is None:
                        continue
                    s = str(val).strip()
                    if not s:
                        continue
                    norm_to_idx[s.casefold()] = c_idx

                if required_norm.issubset(norm_to_idx.keys()):
                    header_map: Dict[str, int] = {}
                    for req in INPUT_REQUIRED_COLUMNS:
                        header_map[req] = norm_to_idx[req.casefold()]
                    return ws, header_map, ws.title, r_idx

        sheet_names = [ws.title for ws in wb.worksheets]
        raise ValueError(
            f"{display_name}: could not find a header row containing required columns in the first "
            f"{HEADER_SCAN_MAX_ROWS} rows of any sheet. Sheets: {sheet_names}"
        )

    def _csv_fieldnames_for_manifest(self) -> List[str]:
        return list(asdict(ManifestRow(
            run_name="",
            role="",
            path="",
            size_bytes=0,
            modified_local="",
            sha256="",
        )).keys())

    def _csv_fieldnames_for_exceptions(self) -> List[str]:
        return list(asdict(ExceptionRow(
            run_name="",
            severity="",
            code="",
            message="",
            source_file="",
            source_sheet="",
            source_row="",
            column="",
            raw_value="",
            action="",
        )).keys())

    def _csv_fieldnames_for_row_evidence(self) -> List[str]:
        return list(asdict(RowEvidence(
            run_name="",
            source_file="",
            source_sheet="",
            source_row=0,
            input_row_hash="",
            processed_flag="",
            output_file="",
            output_sheet="",
            output_row=0,
            output_row_fingerprint_sha256="",
            training_number="",
            training_name="",
            version_raw="",
            revision_norm="",
            trainee_raw="",
            employee_norm="",
            completed_raw="",
            completed_date="",
            status_raw="",
            pass_fail="",
            record_status="",
            rev_zero_stripped="",
            name_reformatted="",
            date_stripped="",
        )).keys())

    # --------- PDF content (fixed-width) ---------

    @staticmethod
    def _kv_block(title: str, items: List[Tuple[str, str]], max_key: int = 24) -> List[str]:
        lines = []
        lines.append(title)
        lines.append("-" * len(title))
        for k, v in items:
            k2 = (k[:max_key]).ljust(max_key)
            lines.append(f"{k2} : {v}")
        lines.append("")
        return lines

    @staticmethod
    def _section(title: str) -> List[str]:
        return [title, "=" * len(title), ""]

    def _build_pdf_lines(
        self,
        input_files: List[Path],
        output_xlsx_files: List[Path],
        manifest_path: Path,
        evidence_path: Path,
        exceptions_path: Path,
        pdf_path: Path,
    ) -> List[str]:
        lines: List[str] = []

        lines.append("TRAINING HISTORY MIGRATION — RUN REPORT")
        lines.append(self.run_name)
        lines.append("")
        # Preface
        lines.extend(self._section("Preface"))
        # Wrap preface later in PDF writer; keep paragraphs as-is
        lines.extend(PDF_PREFACE.split("\n"))
        lines.append("")

        # Run context
        lines.extend(self._section("Run Context"))
        ctx_items = [
            ("Run Name", self.run_name),
            ("Mode", self.context.get("mode", "")),
            ("Input Folder", self.context.get("input_dir", "")),
            ("Output Folder", self.context.get("output_dir", "")),
            ("Operator (OS user)", self.context.get("user", "")),
            ("Machine (hostname)", self.context.get("hostname", "")),
            ("OS", self.context.get("os", "")),
            ("Python", self.context.get("python", "")),
            ("Python executable", self.context.get("python_executable", "")),
            ("Working directory", self.context.get("cwd", "")),
            ("Tool path", self.context.get("tool_path", "")),
            ("openpyxl", self.context.get("openpyxl_version", "")),
            ("reportlab", self.context.get("reportlab_version", "")),
            ("pywin32", self.context.get("pywin32_version", "")),
            ("Started (Local)", self.context.get("started_local", "")),
            ("Started (UTC)", self.context.get("started_utc", "")),
            ("Finished (Local)", self.context.get("finished_local", "")),
            ("Finished (UTC)", self.context.get("finished_utc", "")),
            ("Duration (sec)", str(self.context.get("duration_seconds", ""))),
        ]
        lines.extend(self._kv_block("", ctx_items, max_key=24))

        # Inputs (with SHA pointers via manifest, but keep it readable here)
        lines.extend(self._section("Inputs"))
        for p in input_files:
            lines.append(f"- {p.name}")
            lines.append(f"  {str(p)}")
        lines.append("")

        # Outputs
        lines.extend(self._section("Outputs"))
        for p in output_xlsx_files:
            lines.append(f"- Output XLSX: {p.name}")
            lines.append(f"  {str(p)}")
        lines.append(f"- PDF report: {pdf_path.name}")
        lines.append(f"  {str(pdf_path)}")
        lines.append(f"- CSV manifest: {manifest_path.name}")
        lines.append(f"  {str(manifest_path)}")
        lines.append(f"- CSV row evidence: {evidence_path.name}")
        lines.append(f"  {str(evidence_path)}")
        lines.append(f"- CSV exceptions: {exceptions_path.name}")
        lines.append(f"  {str(exceptions_path)}")
        lines.append("")

        # Per-file verification table
        lines.extend(self._section("Per-File Verification (hashed vs confirmed)"))
        # Table widths in chars will be calculated in writer; here use compact columns.
        headers = ["Source file", "Sheet", "Hdr", "Data", "Scan", "Blank", "Hashed", "Conf", "Out", "P", "F", "Status"]
        rows = []
        for fs in self.file_stats:
            rows.append([
                fs.source_file,
                fs.sheet_used,
                str(fs.header_row),
                str(fs.data_start_row),
                str(fs.rows_scanned_total),
                str(fs.rows_blank_skipped),
                str(fs.eligible_rows_hashed),
                str(fs.hashes_confirmed_processed),
                str(fs.output_rows_written),
                str(fs.pass_count),
                str(fs.fail_count),
                fs.reconciliation_status,
            ])
        # placeholder; widths set at write-time based on max_chars; we’ll render later
        lines.append("__TABLE__PERFILE__")
        # store payload in instance for write-time rendering
        self._pdf_table_perfile = (headers, rows)
        lines.append("")

        # Overall reconciliation
        lines.extend(self._section("Overall Reconciliation"))
        overall_status = "OK"
        if any(fs.reconciliation_status != "OK" for fs in self.file_stats):
            overall_status = "NOT_RECONCILED"
        items = [
            ("Eligible rows hashed (total)", str(self.stats["eligible_rows_hashed_total"])),
            ("Hashes confirmed processed (total)", str(self.stats["hashes_confirmed_processed_total"])),
            ("Blank rows skipped (total)", str(self.stats["rows_blank_skipped_total"])),
            ("Output rows written (total)", str(self.stats["output_rows_written_total"])),
            ("Overall status", overall_status),
            ("Rule", "eligible_hashed = confirmed_processed + intentionally_skipped"),
        ]
        lines.extend(self._kv_block("", items, max_key=30))

        # Data quality metrics
        lines.extend(self._section("Data Quality & Transform Metrics (Overall)"))
        date_min = self.stats["date_min"].strftime("%m/%d/%Y") if self.stats["date_min"] else ""
        date_max = self.stats["date_max"].strftime("%m/%d/%Y") if self.stats["date_max"] else ""
        items = [
            ("Distinct trainees", str(len(self.stats["distinct_trainees"]))),
            ("Distinct courses", str(len(self.stats["distinct_courses"]))),
            ("Completion date range", f"{date_min} → {date_max}" if date_min and date_max else ""),
            ("Pass count", str(self.stats["pass_total"])),
            ("Fail count", str(self.stats["fail_total"])),
            ("Leading-zero revisions stripped", str(self.stats["rev_zero_stripped_count"])),
            ("Names converted First→Last", str(self.stats["name_converted_count"])),
            ("Names already 'Last, First'", str(self.stats["name_already_comma_count"])),
            ("Names single token (WARN)", str(self.stats["name_single_token_count"])),
            ("Dates comma-split (time stripped)", str(self.stats["date_comma_stripped_count"])),
            ("Dates no comma", str(self.stats["date_no_comma_count"])),
            ("Date parse failed (WARN)", str(self.stats["date_parse_failed_count"])),
        ]
        lines.extend(self._kv_block("", items, max_key=32))

        # Status distribution
        lines.extend(self._section("Training Status Distribution (raw)"))
        status_items = sorted(self.stats["status_dist"].items(), key=lambda kv: (-kv[1], kv[0].lower()))
        # Keep top 50 for PDF
        sd_headers = ["status_raw", "count"]
        sd_rows = [[s, str(c)] for s, c in status_items[:50]]
        lines.append("__TABLE__STATUSDIST__")
        self._pdf_table_statusdist = (sd_headers, sd_rows, f"(showing {min(50, len(status_items))} of {len(status_items)})" if len(status_items) > 50 else "")
        lines.append("")

        # Fail summary
        lines.extend(self._section("Fail Summary (by status_raw)"))
        fail_items = [(s, c) for s, c in status_items if not contains_completed(s)]
        fs_headers = ["status_raw", "count"]
        fs_rows = [[s, str(c)] for s, c in fail_items[:50]]
        lines.append("__TABLE__FAILDIST__")
        self._pdf_table_faildist = (fs_headers, fs_rows, f"(showing {min(50, len(fail_items))} of {len(fail_items)})" if len(fail_items) > 50 else "")
        lines.append("")

        # Exceptions summary
        lines.extend(self._section("Exceptions Summary"))
        if self.exceptions:
            by_sev: Dict[str, int] = {}
            by_code: Dict[str, int] = {}
            for e in self.exceptions:
                by_sev[e.severity] = by_sev.get(e.severity, 0) + 1
                by_code[e.code] = by_code.get(e.code, 0) + 1

            lines.append("Counts by severity:")
            for s, c in sorted(by_sev.items(), key=lambda kv: (-kv[1], kv[0])):
                lines.append(f"  {s:>5} : {c}")
            lines.append("")
            lines.append("Counts by code:")
            for c, n in sorted(by_code.items(), key=lambda kv: (-kv[1], kv[0])):
                lines.append(f"  {c:<28} : {n}")
            lines.append("")
            lines.append("Examples (first 15):")
            for e in self.exceptions[:15]:
                lines.append(f"- {e.severity}/{e.code} @ {e.source_file}#{e.source_row} [{e.column}]")
                lines.extend(FixedWidthPDF.wrap_text(f"  {e.message} (action={e.action})", 200, indent=2))
            lines.append("")
            lines.append("Full exception detail is provided in the exceptions CSV.")
        else:
            lines.append("No exceptions recorded.")
        lines.append("")

        # Evidence pointer
        lines.extend(self._section("Row-Level Evidence"))
        lines.append(f"Row evidence chain is provided in: {evidence_path.name}")
        lines.append("This file maps each eligible input-row hash to its confirmed processing and resulting output-row hash.")
        lines.append("")

        return lines

    def _render_embedded_tables(self, lines: List[str], max_chars: int) -> List[str]:
        """
        Replace __TABLE__... placeholders with fixed-width tables sized to max_chars.
        """
        out: List[str] = []
        for ln in lines:
            if ln == "__TABLE__PERFILE__":
                headers, rows = self._pdf_table_perfile
                # widths sum must be <= max_chars; choose compact widths
                # total separators: 3*(n-1). n=12 => 33 chars separators.
                # allocate remaining to columns.
                col_widths = [16, 12, 3, 4, 5, 5, 6, 5, 4, 2, 2, 9]
                # if page is narrower, shrink the first columns
                total_needed = sum(col_widths) + 3 * (len(col_widths) - 1)
                if total_needed > max_chars:
                    shrink = total_needed - max_chars
                    # shrink Source file then Sheet
                    s1 = min(shrink, max(0, col_widths[0] - 10))
                    col_widths[0] -= s1
                    shrink -= s1
                    s2 = min(shrink, max(0, col_widths[1] - 8))
                    col_widths[1] -= s2
                table_lines = FixedWidthPDF.make_table(headers, rows, col_widths)
                out.extend(table_lines)
                continue

            if ln == "__TABLE__STATUSDIST__":
                headers, rows, note = self._pdf_table_statusdist
                col_widths = [max(20, min(max_chars - (3 + 8), 70)), 8]
                table_lines = FixedWidthPDF.make_table(headers, rows, col_widths)
                out.extend(table_lines)
                if note:
                    out.append(note)
                continue

            if ln == "__TABLE__FAILDIST__":
                headers, rows, note = self._pdf_table_faildist
                if not rows:
                    out.append("No fail statuses.")
                    continue
                col_widths = [max(20, min(max_chars - (3 + 8), 70)), 8]
                table_lines = FixedWidthPDF.make_table(headers, rows, col_widths)
                out.extend(table_lines)
                if note:
                    out.append(note)
                continue

            out.append(ln)
        return out

    # ------------- Main run -------------

    def run(self, input_paths: List[Path], mode: str):
        """
        mode: "per_file" or "consolidated"
        """
        self.run_dir.mkdir(parents=True, exist_ok=True)

        self._log(f"[run] Run Name: {self.run_name}")
        self._log(f"[run] Run folder: {self.run_dir}")

        # Hash original inputs and add to manifest as INPUT
        for p in input_paths:
            self._check_pause_cancel()
            self._log(f"[hash] SHA-256 input: {p.name}")
            st = p.stat()
            self.manifest.append(ManifestRow(
                run_name=self.run_name,
                role="INPUT",
                path=str(p),
                size_bytes=int(st.st_size),
                modified_local=datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                sha256=sha256_file(p),
            ))

        output_xlsx_files: List[Path] = []
        consolidated_rows: List[List[str]] = []
        consolidated_output_path = self.run_dir / f"{self.run_name}_output.xlsx"
        consolidated_output_row_cursor = 2  # output data starts at row 2

        for file_idx, original_path in enumerate(input_paths, start=1):
            self._check_pause_cancel()

            xlsx_path, display_name, converted_path = self._normalize_input_to_xlsx(original_path)
            src_type = original_path.suffix.lower().lstrip(".")
            self._log(f"[read] Loading: {display_name}")

            if mode == "per_file":
                out_file_name = f"{self.run_name}_input{file_idx:02d}_{safe_filename(original_path.stem)}_output.xlsx"
                out_path = self.run_dir / out_file_name
            else:
                out_file_name = consolidated_output_path.name
                out_path = consolidated_output_path

            wb = load_workbook(xlsx_path, read_only=True, data_only=True)

            # Per-file stats
            file_rows_scanned_total = 0
            file_blank_skipped = 0
            file_eligible_hashed = 0
            file_processed_confirmed = 0
            file_output_rows_written = 0
            file_pass = 0
            file_fail = 0
            file_distinct_trainees: set = set()
            file_distinct_courses: set = set()
            file_date_min: Optional[datetime] = None
            file_date_max: Optional[datetime] = None
            sheet_used = ""
            header_row_idx = 0
            data_start_row = 0

            try:
                ws, header_map, sheet_used, header_row_idx = self._detect_header_in_workbook(wb, display_name)
                data_start_row = header_row_idx + 1
                self._log(f"[read] Using sheet '{sheet_used}' header row {header_row_idx} (data starts row {data_start_row})")

                # verification sets key
                key = (display_name, sheet_used)
                self._eligible_hashes_by_key.setdefault(key, set())
                self._processed_hashes_by_key.setdefault(key, set())

                out_rows_for_this_file: List[List[str]] = []
                per_file_output_row_cursor = 2

                for excel_row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                    self._check_pause_cancel()
                    file_rows_scanned_total += 1

                    def get(col: str) -> str:
                        v = row[header_map[col]] if row and header_map[col] < len(row) else None
                        return "" if v is None else str(v).strip()

                    training_number = get("Training Number")
                    training_name = get("Training Name")
                    version_raw = get("Version")
                    trainee_raw = get("Trainee Name")
                    completed_raw = get("Completed Date/Time")
                    status_raw = get("Training Status")

                    # Skip fully blank “data” rows (not eligible for hashing)
                    if not any([training_number, training_name, version_raw, trainee_raw, completed_raw, status_raw]):
                        file_blank_skipped += 1
                        self.stats["rows_blank_skipped_total"] += 1
                        continue

                    # Eligible row: hash (catalog)
                    ih = input_row_hash(
                        source_file=display_name,
                        source_sheet=sheet_used,
                        source_row=excel_row_idx,
                        training_number=training_number,
                        training_name=training_name,
                        version_raw=version_raw,
                        trainee_raw=trainee_raw,
                        completed_raw=completed_raw,
                        status_raw=status_raw,
                    )
                    self._eligible_hashes_by_key[key].add(ih)
                    self.stats["eligible_rows_hashed_total"] += 1
                    file_eligible_hashed += 1

                    # Create evidence record (processed_flag N initially)
                    if ih not in self._evidence_by_hash:
                        self._evidence_order.append(ih)
                        self._evidence_by_hash[ih] = RowEvidence(
                            run_name=self.run_name,
                            source_file=display_name,
                            source_sheet=sheet_used,
                            source_row=excel_row_idx,
                            input_row_hash=ih,
                            processed_flag="N",
                            output_file="",
                            output_sheet="",
                            output_row=0,
                            output_row_fingerprint_sha256="",
                            training_number=training_number,
                            training_name=training_name,
                            version_raw=version_raw,
                            revision_norm="",
                            trainee_raw=trainee_raw,
                            employee_norm="",
                            completed_raw=completed_raw,
                            completed_date="",
                            status_raw=status_raw,
                            pass_fail="",
                            record_status="",
                            rev_zero_stripped="N",
                            name_reformatted="N",
                            date_stripped="N",
                        )

                    # Process row
                    revision_norm = strip_leading_zeroes(version_raw)
                    rev_zero_stripped_flag = "Y" if (
                        str(version_raw).strip().isdigit()
                        and str(version_raw).strip().startswith("0")
                        and revision_norm != str(version_raw).strip()
                    ) else "N"
                    if rev_zero_stripped_flag == "Y":
                        self.stats["rev_zero_stripped_count"] += 1

                    employee_norm, name_applied, name_note = name_first_last_to_last_first(trainee_raw)
                    if name_applied:
                        self.stats["name_converted_count"] += 1
                    elif name_note == "already_comma":
                        self.stats["name_already_comma_count"] += 1
                    elif name_note == "single_token":
                        self.stats["name_single_token_count"] += 1
                        self._add_exception(
                            severity="WARN",
                            code="NAME_SINGLE_TOKEN",
                            message="Trainee name could not be confidently reformatted (single token). Left as-is.",
                            source_file=display_name,
                            source_sheet=sheet_used,
                            source_row=str(excel_row_idx),
                            column="Trainee Name",
                            raw_value=trainee_raw,
                            action="left_as_is",
                        )

                    completed_date, date_stripped, date_note = strip_date_only(completed_raw)
                    if date_note == "comma_split":
                        self.stats["date_comma_stripped_count"] += 1
                    else:
                        self.stats["date_no_comma_count"] += 1

                    dt_parsed = self._parse_date_for_range(completed_date)
                    if dt_parsed:
                        if self.stats["date_min"] is None or dt_parsed < self.stats["date_min"]:
                            self.stats["date_min"] = dt_parsed
                        if self.stats["date_max"] is None or dt_parsed > self.stats["date_max"]:
                            self.stats["date_max"] = dt_parsed
                        if file_date_min is None or dt_parsed < file_date_min:
                            file_date_min = dt_parsed
                        if file_date_max is None or dt_parsed > file_date_max:
                            file_date_max = dt_parsed
                    else:
                        if completed_date.strip():
                            self.stats["date_parse_failed_count"] += 1
                            self._add_exception(
                                severity="WARN",
                                code="DATE_PARSE_FAILED",
                                message="Completion date did not match expected date format; wrote stripped token as-is.",
                                source_file=display_name,
                                source_sheet=sheet_used,
                                source_row=str(excel_row_idx),
                                column="Completed Date/Time",
                                raw_value=completed_raw,
                                action="kept_date_token",
                            )

                    # Status mapping + distribution
                    self.stats["status_dist"][status_raw] = self.stats["status_dist"].get(status_raw, 0) + 1
                    completed_flag = contains_completed(status_raw)
                    if completed_flag:
                        pass_fail = "Pass"
                        record_status = "Passed"
                        file_pass += 1
                        self.stats["pass_total"] += 1
                    else:
                        pass_fail = "Fail"
                        record_status = "Failed"
                        file_fail += 1
                        self.stats["fail_total"] += 1

                    out_row = [
                        training_name,
                        revision_norm,
                        "",
                        employee_norm,
                        "",
                        completed_date,
                        pass_fail,
                        record_status,
                        "",
                        "",
                    ]
                    oh = output_row_hash(out_row)

                    # Output row numbering
                    if mode == "consolidated":
                        out_row_num = consolidated_output_row_cursor
                        consolidated_output_row_cursor += 1
                        consolidated_rows.append(out_row)
                    else:
                        out_row_num = per_file_output_row_cursor
                        per_file_output_row_cursor += 1
                        out_rows_for_this_file.append(out_row)

                    file_output_rows_written += 1
                    self.stats["output_rows_written_total"] += 1

                    # distinct sets
                    if employee_norm:
                        self.stats["distinct_trainees"].add(employee_norm)
                        file_distinct_trainees.add(employee_norm)
                    if training_name:
                        self.stats["distinct_courses"].add(training_name)
                        file_distinct_courses.add(training_name)

                    # confirm processed for verification
                    self._processed_hashes_by_key[key].add(ih)
                    file_processed_confirmed += 1
                    self.stats["hashes_confirmed_processed_total"] += 1

                    # Update evidence record with processed outputs + hashes
                    ev = self._evidence_by_hash[ih]
                    ev.processed_flag = "Y"
                    ev.output_file = out_file_name
                    ev.output_sheet = "Training Migration"
                    ev.output_row = out_row_num
                    ev.output_row_fingerprint_sha256 = oh
                    ev.revision_norm = revision_norm
                    ev.employee_norm = employee_norm
                    ev.completed_date = completed_date
                    ev.pass_fail = pass_fail
                    ev.record_status = record_status
                    ev.rev_zero_stripped = rev_zero_stripped_flag
                    ev.name_reformatted = "Y" if name_applied else "N"
                    ev.date_stripped = "Y" if date_stripped else "N"
                    self._evidence_by_hash[ih] = ev

                # write per-file output
                if mode == "per_file":
                    self._log(f"[write] Output XLSX: {out_path.name}")
                    self._write_output_xlsx(out_path, out_rows_for_this_file)
                    output_xlsx_files.append(out_path)

            finally:
                wb.close()

            # reconcile per file/sheet
            key = (display_name, sheet_used)
            eligible_set = self._eligible_hashes_by_key.get(key, set())
            processed_set = self._processed_hashes_by_key.get(key, set())
            unaccounted = eligible_set - processed_set
            reconciliation_status = "OK" if len(unaccounted) == 0 else "NOT_RECONCILED"
            if len(unaccounted) > 0:
                self._add_exception(
                    severity="ERROR",
                    code="VERIFICATION_MISMATCH",
                    message=f"{len(unaccounted)} eligible input hashes were not confirmed processed.",
                    source_file=display_name,
                    source_sheet=sheet_used,
                    source_row="",
                    column="",
                    raw_value="",
                    action="review_row_evidence",
                )

            self.file_stats.append(FileRunStats(
                source_file=display_name,
                source_path=str(original_path),
                source_type=src_type,
                converted_path=converted_path,
                sheet_used=sheet_used,
                header_row=header_row_idx,
                data_start_row=data_start_row,
                rows_scanned_total=file_rows_scanned_total,
                rows_blank_skipped=file_blank_skipped,
                eligible_rows_hashed=file_eligible_hashed,
                hashes_confirmed_processed=file_processed_confirmed,
                unaccounted_rows=len(unaccounted),
                output_rows_written=file_output_rows_written,
                pass_count=file_pass,
                fail_count=file_fail,
                distinct_trainees=len(file_distinct_trainees),
                distinct_courses=len(file_distinct_courses),
                date_min=file_date_min.strftime("%m/%d/%Y") if file_date_min else "",
                date_max=file_date_max.strftime("%m/%d/%Y") if file_date_max else "",
                reconciliation_status=reconciliation_status,
            ))

        # consolidated output write
        if mode == "consolidated":
            self._check_pause_cancel()
            self._log(f"[write] Output XLSX: {consolidated_output_path.name}")
            self._write_output_xlsx(consolidated_output_path, consolidated_rows)
            output_xlsx_files.append(consolidated_output_path)

        # finalize run times
        self.run_duration_seconds = int(time.time() - self.run_started_ts)
        self.run_finished_local = datetime.now().strftime(f"%m/%d/%Y %I:%M:%S %p {LOCAL_TZ_LABEL}")
        self.run_finished_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        self.context["finished_local"] = self.run_finished_local
        self.context["finished_utc"] = self.run_finished_utc
        self.context["duration_seconds"] = self.run_duration_seconds

        # output paths
        manifest_path = self.run_dir / f"{self.run_name}_run_manifest.csv"
        evidence_path = self.run_dir / f"{self.run_name}_row_evidence.csv"
        exceptions_path = self.run_dir / f"{self.run_name}_exceptions.csv"
        pdf_path = self.run_dir / f"{self.run_name}_run_report.pdf"

        # write CSV logs
        self._check_pause_cancel()
        self._log("[write] Writing audit CSV logs")

        # Row evidence (stable order)
        evidence_rows = [asdict(self._evidence_by_hash[h]) for h in self._evidence_order]
        self._write_csv(evidence_path, self._csv_fieldnames_for_row_evidence(), evidence_rows)

        self._write_csv(exceptions_path, self._csv_fieldnames_for_exceptions(), [asdict(r) for r in self.exceptions])

        # hash outputs + logs (except manifest)
        for out_xlsx in output_xlsx_files:
            self._check_pause_cancel()
            self._log(f"[hash] SHA-256 output: {out_xlsx.name}")
            self._manifest_add_file("OUTPUT", out_xlsx)

        self._log(f"[hash] SHA-256 log: {evidence_path.name}")
        self._manifest_add_file("LOG", evidence_path)
        self._log(f"[hash] SHA-256 log: {exceptions_path.name}")
        self._manifest_add_file("LOG", exceptions_path)

        # build PDF (fixed width)
        self._check_pause_cancel()
        self._log("[write] Building PDF run report")

        pdf_writer = FixedWidthPDF(pdf_path=pdf_path, run_name=self.run_name, font_name="Courier", font_size=9, leading=11)
        max_chars = pdf_writer._compute_max_chars()

        pdf_lines = self._build_pdf_lines(
            input_files=input_paths,
            output_xlsx_files=output_xlsx_files,
            manifest_path=manifest_path,
            evidence_path=evidence_path,
            exceptions_path=exceptions_path,
            pdf_path=pdf_path,
        )
        pdf_lines = self._render_embedded_tables(pdf_lines, max_chars=max_chars)
        pdf_writer.write(pdf_lines)

        # hash report and record
        self._log(f"[hash] SHA-256 report: {pdf_path.name}")
        self._manifest_add_file("REPORT", pdf_path)

        # write manifest last (do not self-hash inside itself)
        self._write_csv(manifest_path, self._csv_fieldnames_for_manifest(), [asdict(r) for r in self.manifest])

        self._log("[done] Run complete.")
        return {
            "run_name": self.run_name,
            "run_dir": str(self.run_dir),
            "outputs_xlsx": [str(p) for p in output_xlsx_files],
            "pdf": str(pdf_path),
            "csvs": [str(manifest_path), str(evidence_path), str(exceptions_path)],
        }


# =========================
# File discovery
# =========================

def discover_excel_files(parent_folder: Path) -> List[Path]:
    if not parent_folder.exists() or not parent_folder.is_dir():
        raise FileNotFoundError(str(parent_folder))

    files: List[Path] = []
    for p in parent_folder.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() in (".xls", ".xlsx"):
            files.append(p)

    files.sort(key=lambda x: x.name.lower())
    return files


# =========================
# Tkinter GUI
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Training History Migration Builder")
        self.geometry("980x680")

        self.log_q: queue.Queue = queue.Queue()
        self.cancel_evt = threading.Event()
        self.pause_evt = threading.Event()
        self.worker_thread: Optional[threading.Thread] = None

        self.input_dir_var = tk.StringVar(value="")
        self.output_dir_var = tk.StringVar(value="")
        self.mode_var = tk.StringVar(value="consolidated")
        self.discovered_var = tk.StringVar(value="No input folder selected.")

        self._build_ui()
        self.after(100, self._drain_log_queue)

    def _build_ui(self):
        root = ttk.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        in_box = ttk.LabelFrame(root, text="Input (Parent Folder)", padding=10)
        in_box.pack(fill="x", padx=4, pady=6)

        in_row = ttk.Frame(in_box)
        in_row.pack(fill="x", pady=4)
        ttk.Label(in_row, text="Input Folder:", width=14).pack(side="left")
        ttk.Entry(in_row, textvariable=self.input_dir_var).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(in_row, text="Browse…", command=self._browse_input_dir).pack(side="left")

        ttk.Label(in_box, textvariable=self.discovered_var).pack(anchor="w", pady=(6, 0))

        out_box = ttk.LabelFrame(root, text="Output", padding=10)
        out_box.pack(fill="x", padx=4, pady=6)

        out_row = ttk.Frame(out_box)
        out_row.pack(fill="x", pady=4)
        ttk.Label(out_row, text="Output Folder:", width=14).pack(side="left")
        ttk.Entry(out_row, textvariable=self.output_dir_var).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(out_row, text="Browse…", command=self._browse_output_dir).pack(side="left")

        mode_row = ttk.Frame(out_box)
        mode_row.pack(fill="x", pady=8)
        ttk.Label(mode_row, text="Output Mode:", width=14).pack(side="left")
        ttk.Radiobutton(
            mode_row,
            text="One output file for ALL input files (consolidated)",
            value="consolidated",
            variable=self.mode_var,
        ).pack(side="left", padx=6)
        ttk.Radiobutton(
            mode_row,
            text="One output file PER input file",
            value="per_file",
            variable=self.mode_var,
        ).pack(side="left", padx=14)

        ctl = ttk.Frame(root)
        ctl.pack(fill="x", padx=4, pady=10)

        self.run_btn = ttk.Button(ctl, text="Run", command=self._on_run)
        self.run_btn.pack(side="left")

        self.pause_btn = ttk.Button(ctl, text="Pause", command=self._on_pause_toggle, state="disabled")
        self.pause_btn.pack(side="left", padx=8)

        self.cancel_btn = ttk.Button(ctl, text="Cancel", command=self._on_cancel, state="disabled")
        self.cancel_btn.pack(side="left")

        ttk.Separator(root).pack(fill="x", padx=4, pady=8)

        logs_box = ttk.LabelFrame(root, text="Run Log", padding=10)
        logs_box.pack(fill="both", expand=True, padx=4, pady=6)

        self.log_text = tk.Text(logs_box, height=18, wrap="word")
        self.log_text.pack(fill="both", expand=True)
        self.log_text.configure(state="disabled")

    def _browse_input_dir(self):
        path = filedialog.askdirectory(title="Select Input Parent Folder")
        if path:
            self.input_dir_var.set(path)
            self._refresh_discovered()

    def _browse_output_dir(self):
        path = filedialog.askdirectory(title="Select Output Folder")
        if path:
            self.output_dir_var.set(path)

    def _refresh_discovered(self):
        p = self.input_dir_var.get().strip()
        if not p:
            self.discovered_var.set("No input folder selected.")
            return
        try:
            files = discover_excel_files(Path(p))
            preview = ", ".join([f.name for f in files[:6]])
            if len(files) > 6:
                preview += " ..."
            self.discovered_var.set(f"Discovered {len(files)} Excel file(s): {preview}")
        except Exception as e:
            self.discovered_var.set(f"Discovery error: {e}")

    def _append_log(self, msg: str):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _drain_log_queue(self):
        try:
            while True:
                kind, msg = self.log_q.get_nowait()
                if kind == "log":
                    self._append_log(msg)
        except queue.Empty:
            pass
        self.after(100, self._drain_log_queue)

    def _set_running_state(self, running: bool):
        self.run_btn.configure(state="disabled" if running else "normal")
        self.pause_btn.configure(state="normal" if running else "disabled")
        self.cancel_btn.configure(state="normal" if running else "disabled")

    def _on_pause_toggle(self):
        if not self.worker_thread or not self.worker_thread.is_alive():
            return
        if self.pause_evt.is_set():
            self.pause_evt.clear()
            self.pause_btn.configure(text="Pause")
            self._append_log("[ui] Resumed.")
        else:
            self.pause_evt.set()
            self.pause_btn.configure(text="Resume")
            self._append_log("[ui] Paused.")

    def _on_cancel(self):
        if self.worker_thread and self.worker_thread.is_alive():
            self.cancel_evt.set()
            self._append_log("[ui] Cancel requested...")

    def _on_run(self):
        input_dir = self.input_dir_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        mode = self.mode_var.get().strip()

        if not input_dir:
            messagebox.showerror("Missing input folder", "Please select an input parent folder.")
            return
        if not output_dir:
            messagebox.showerror("Missing output folder", "Please select an output folder.")
            return

        try:
            input_paths = discover_excel_files(Path(input_dir))
        except Exception as e:
            messagebox.showerror("Input discovery error", str(e))
            return

        if not input_paths:
            messagebox.showerror("No input files", "No .xls/.xlsx files found in the selected input folder.")
            return

        self.cancel_evt.clear()
        self.pause_evt.clear()
        self.pause_btn.configure(text="Pause")
        self._set_running_state(True)
        self._append_log("[ui] Starting run...")
        self._append_log(f"[ui] Found {len(input_paths)} input file(s). Mode={mode}")

        run_name = f"{RUN_PREFIX}_{datetime.now().strftime('%m%d%Y_%H%M%S')}"
        run_dir = Path(output_dir) / run_name

        tool_path = ""
        try:
            tool_path = str(Path(__file__).resolve())
        except Exception:
            tool_path = ""

        def worker():
            try:
                tx = TrainingTransformer(
                    run_name=run_name,
                    run_dir=run_dir,
                    log_q=self.log_q,
                    cancel_evt=self.cancel_evt,
                    pause_evt=self.pause_evt,
                    input_dir=input_dir,
                    output_dir=output_dir,
                    mode=mode,
                    tool_path=tool_path,
                )
                result = tx.run(input_paths=input_paths, mode=mode)

                self.log_q.put(("log", f"[result] Run folder: {result['run_dir']}"))
                for p in result["outputs_xlsx"]:
                    self.log_q.put(("log", f"[result] Output XLSX: {p}"))
                self.log_q.put(("log", f"[result] PDF report: {result['pdf']}"))
                for p in result["csvs"]:
                    self.log_q.put(("log", f"[result] CSV log: {p}"))
                self.log_q.put(("log", "[ui] Completed successfully."))
            except Exception as e:
                self.log_q.put(("log", f"[error] {type(e).__name__}: {e}"))
            finally:
                self.after(0, lambda: self._set_running_state(False))

        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()


if __name__ == "__main__":
    App().mainloop()
